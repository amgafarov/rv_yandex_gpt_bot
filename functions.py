import os.path
import textwrap
import docx2txt
import striprtf.striprtf as rtf_reader
import pdfminer.high_level
import logging
import openpyxl
from datetime import datetime
from telegram_secure import __admins__ as admins
from PIL import Image, ImageDraw, ImageFont
import rembg
import base64
from yandex_services.yandex_gpt_request import get_text_from_image_id, get_text_from_image_by_id, __secret_key_speech__
import time
import pillow_heif
from moviepy.editor import *
from speechkit import model_repository, configure_credentials, creds


# TODO : audio to text func
# def transcribe_audio(path: str):
#    print('HERE')
#    r = speech_recognition.Recognizer()
#    with speech_recognition.AudioFile(path) as audio_file:
#        print('START')
#       try:
#            text = r.recognize_google_cloud(audio_file)
#            print('Text:\n' + text)
#        except Exception as e:
#            print('Exception:\n' + str(e.with_traceback))


def text_from_pdf_scan(pdf_file: str):
    try:
        img_type = 'application/pdf'
        with open(pdf_file, 'rb') as img:
            content = img.read()

        response = get_text_from_image_id(base64_image=base64.b64encode(content).decode('utf-8'), image_type=img_type)
        id = response['id']
        while True:
            ans = get_text_from_image_by_id(str(id))
            if 'error' in ans[0]:
                time.sleep(2)
                continue
            break

        res = ''
        for i in ans:
            for alt in i['result']['textAnnotation']['blocks']:
                for line in alt['lines']:
                    res += line['text']
                    res += '\n'
    except Exception as e:
        raise e
    finally:
        if os.path.exists(pdf_file):
            os.remove(pdf_file)
    return res


def convert_heic_to_jpg(file_path: str) -> str:
    pillow_heif.register_heif_opener()
    new_path = file_path + '.jpg'
    image = Image.open(file_path)
    image.save(new_path, format='JPEG',
               quality=90, optimize=True, progressive=True)
    return new_path


def get_users() -> []:
    logging.info(f'{datetime.now()} GET USERS')
    f = open('accepted_users', 'r')
    content = f.read()
    f.close()
    return content.split(',')


def read_text_file(src):
    text = ''
    if str(src).endswith('.docx'):
        try:
            text = docx2txt.process(src)
        except Exception as e:
            raise e
    elif src.endswith('.txt'):
        try:
            with open(src) as f:
                text = f.read()
        except Exception as e:
            raise e
    # TODO : .doc reader
    elif src.endswith('.doc'):
        try:
            raise Exception
        except Exception as e:
            raise e
    elif src.endswith('.rtf'):
        try:
            with open(src) as rtf_file:
                content = rtf_file.read()
                text = rtf_reader.rtf_to_text(content)
        except Exception as e:
            raise e
    elif src.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(src)
            sheet_names = wb.sheetnames
            ws = wb[sheet_names[0]]
            for row in ws.iter_rows(min_row=1, values_only=True):
                try:
                    text += "\t".join(str(row)) + "\n"
                except TypeError:
                    continue
        except Exception as e:
            raise e
    elif src.endswith('.pdf'):
        text = text_from_pdf_scan(src)

    return text


def add_user(user_id=str):
    user_id = str(user_id)
    if len(user_id) == 0 or user_id is None:
        return
    try:
        int(user_id)
    except:
        logging.error(f'INVALID USER ID {user_id}')
        return

    logging.info(f'{datetime.now()} ADD USER {user_id}')
    with open('accepted_users', 'a') as f:
        f.write(f'{str(user_id)},')
        logging.info(f'USER ID ADDED : {user_id}')


def get_secure(user_id) -> bool:
    allowed_users = get_users()
    if str(user_id) not in admins and str(user_id) not in allowed_users:
        return False
    else:
        return True


def create_image(text: str, user_id: str) -> str:
    img = 'image_creator/telegram_image.png'
    created_image = f'image_creator/{user_id}-telegram_image.png'
    font = ImageFont.truetype("image_creator/ofont.ru_Helios.ttf", 110)

    image = Image.open(img)
    drawer = ImageDraw.Draw(image)
    lines = textwrap.wrap(text, width=15)
    text_for_print = ''

    for line in lines:
        text_for_print += line.upper().strip() + '\n'

    drawer.multiline_text((image.width / 2, image.height / 2), text_for_print.strip(),
                          font=font, fill='white', align='center',
                          anchor='mm')

    image.save(created_image)

    return created_image


def create_m_image(text: str, user_id: str) -> str:
    img = 'image_creator/telegram_image.png'
    created_image = f'image_creator/{user_id}-telegram_image.png'
    font = ImageFont.truetype("image_creator/ofont.ru_Helios.ttf", 110)
    image = Image.open(img)
    drawer = ImageDraw.Draw(image)
    drawer.multiline_text((image.width / 2, image.height / 2), text.upper().strip(),
                          font=font, fill='white', align='center',
                          anchor='mm')
    image.save(created_image)

    return created_image


def remove_image_background(image: str):
    created_image = f'{image}-output.png'
    input_file = Image.open(image)
    output = rembg.remove(input_file)
    output.save(created_image)
    return created_image


def image_watermark(image_path: str) -> str:
    image = Image.open(image_path)
    new_image_path = f'files/{image_path.split("/")[-1].split(".")[0]}-WM.{image_path.split(".")[-1]}'
    watermark = Image.open('image_creator/watermark.png')
    watermark_mask = Image.open('image_creator/watermark_mask_80.png').convert('L')

    if image.width > image.height:
        k = 3  # Коэффициент уменьшения водяного знака
        step = 50  # Отступ от краев изображения
    else:
        k = 2
        step = 25

    scale = image.width / (k * watermark.width)  # Как сильно изменять размер водяного знака

    wm_width = int(watermark.width * scale)  # Новая ширина водяного знака
    wm_height = int(watermark.height * scale)  # Новая высота водяного знака

    watermark = watermark.resize((wm_width, wm_height))  # Изменение размера водяного знака
    watermark_mask = watermark_mask.resize((wm_width, wm_height))  # Изменение размера маски водяного знака

    position_x = image.width - step - wm_width  # Позиция по ширине для водяного знака
    position_y = image.height - step - wm_height  # Позиция по высоте для водяного знака

    image.paste(watermark, (position_x, position_y), mask=watermark_mask)  # Нанесение водяного знака
    image.save(new_image_path)
    image.close()
    watermark.close()
    watermark_mask.close()
    return new_image_path


def video_watermark(video_path: str) -> str:
    new_file = video_path + '-WM.mp4'
    wm_image = 'image_creator/watermark_video.png'

    video1 = VideoFileClip(video_path)
    video_width = video1.w
    video_height = video1.h
    step = 50

    watermark = Image.open(wm_image)

    if video_width >= video_height:  # Горизонтальное
        k = (video_width - 5 * step) / (4 * watermark.width)
        watermark = watermark.resize((int(watermark.width * k), int(watermark.height * k)))

    else:  # Вертикальное
        k = (video_width - 2 * step) / watermark.width
        watermark = watermark.resize((int(watermark.width * k), int(watermark.height * k)))

    pos_y = step

    resized_wm = 'image_creator/resized_wm.png'
    watermark.save(resized_wm)

    elements = [video1]

    while pos_y + watermark.height < video_height:
        pos_x = step
        while pos_x + watermark.width < video_width:
            elements.append(ImageClip(resized_wm)
                            .with_position((pos_x, pos_y))
                            .with_duration(video1.duration))
            pos_x += watermark.width + step

        pos_y += step + watermark.height

    final = CompositeVideoClip(elements)
    final.write_videofile(new_file, audio=True, codec="libx264", threads=4)

    if os.path.exists(resized_wm):
        os.remove(resized_wm)

    return new_file


def text_to_speech(text: str, user_id: str) -> str:
    export_path_wav = 'files/' + user_id + '.wav'

    configure_credentials(
        yandex_credentials=creds.YandexCredentials(
            api_key=__secret_key_speech__
        )
    )

    model = model_repository.synthesis_model()
    model.voice = 'anton'
    #model.role = 'good'

    result = model.synthesize(text, raw_format=False)
    result.export(export_path_wav, 'wav')

    return export_path_wav

